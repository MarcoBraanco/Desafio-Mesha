import openpyxl
from datetime import datetime
import os

class ExcelFunctions:
    def __init__(self):
        self.columnNames = {1:"ID", 2: "Invoice Number", 3:"invoice Date", 4:"First Name", 5:"Last Name",
                6:"Company Name", 7:"Role in Company", 8:"Email", 9:"Address", 10: "Phone Number", 11: "Status"}
        self.currentPath = os.getcwd()
        print("teste")
        
        
    def CreateExcel(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Invoice Data"

    
    def WriteExcel(self, data):
        self.CreateExcel()

        headerRow = [self.columnNames[col] for col in range(1, len(self.columnNames) + 1)]
        self.ws.append(headerRow)


        for chave, valor in data.items():
            self.ws.append([chave] + valor)


        nowTimeStamp = datetime.now().strftime("%d.%m.%Y - %Hh %Mm %Ss")
        filePath = f"{self.currentPath}\\Data\\Invoice extraction data - {nowTimeStamp}.xlsx"
        self.wb.save(filePath)

        return filePath


    def ReadExcel(self, file_path):

        self.wb = openpyxl.load_workbook(file_path)

        self.ws = self.wb.active
        data = {}
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            chave = row[0]
            valores = list(row[1:])
            data[chave] = valores

        return data
    
    def UpdateStatusExcel(self, id, filePath, status):
        try:

            workbook = openpyxl.load_workbook(filePath)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] == id:
                    coluna_status = 11 #Coluna status
                    sheet.cell(row=int(row[0]) + 1, column=coluna_status, value=status)
                    break

            workbook.save(filePath)

        except Exception as e:
            print(f"Erro ao atualizar o status do item: {e}")


    def CheckStatus(self, filePath):

        workbook = openpyxl.load_workbook(filePath)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, max_col=11, values_only=True):
            if row[10] == None:
                sheet.cell(row=int(row[0]) + 1, column = 11, value="Erro: Tratar manualmente")
        
        workbook.save(filePath)


    def WriteLog(self, mensagem, filePath):

        with open(filePath, "a") as arquivo:
            nowTimeStamp = datetime.now().strftime("%d/%m/%Y - %Hh %Mm %Ss")
            arquivo.write(f"{nowTimeStamp} - {mensagem}\n")
